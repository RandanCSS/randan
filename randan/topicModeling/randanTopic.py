#!/usr/bin/env python
# coding: utf-8

'''
(EN) A module for topic modelling
(RU) Модуль для тематического моделирования
'''

# 0 Активировать требуемые для работы скрипта модули и пакеты + пререквизиты

# sys & subprocess -- эти пакеты должны быть предустановлены. Если с ними какая-то проблема, то из этого скрипта решить их сложно
import sys
from subprocess import check_call

# --- остальные модули и пакеты
attempt = 0
while True:
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.utils.dataframe import dataframe_to_rows
        from randan import descriptive_statistics, dimension_reduction
        from tqdm import tqdm
        import matplotlib.pyplot as plt, os, pandas, time, warnings
        break
    except ModuleNotFoundError:
        errorDescription = sys.exc_info()
        module = str(errorDescription[1]).replace("No module named '", '').replace("'", '') #.replace('_', '')
        if '.' in module: module = module.split('.')[1]
        print(
f'''Пакет {module} НЕ прединсталлирован, но он требуется для работы скрипта, поэтому будет инсталлирован сейчас
Попытка № {attempt} из 10
'''
              )
        check_call([sys.executable, "-m", "pip", "install", module])
        attempt += 1
        if  attempt == 10:
            print(
f'''Пакет {module} НЕ прединсталлирован; он требуется для работы скрипта, но инсталлировать его не удаётся,
поэтому попробуйте инсталлировать его вручную, после чего снова запустите скрипт
'''
                  )
            break
tqdm.pandas() # для визуализации прогресса функций, применяемых к датафреймам

# 1 Авторские функции для..
# 1.0 ..проверки наличия в поданном пользователем датафрейме ключевого тестового столбца
def columnTargetChecker(columnTarget, columnTargetMessage, df):
    if columnTarget not in df.columns:
        print(f"В поданном Вами датафрейме отсутствует столбец '{columnTarget}' .", columnTargetMessage)
        while True:
            print('--- Впишите, пожалуйста, его название')
            columnTarget = input()
            if columnTarget in df.columns: break
            else:
                print('--- Вы вписали что-то не то')
    return columnTarget

# 1.1 ..обработки документов полюса топика с учётом дублирующихся тестов
def docsSelector(dfUnique_topicScoreS, minusPlus, docsLimit, topicName):
    dfUnique_topicScoreS = dfUnique_topicScoreS.sort_values(topicName, ascending=False) if minusPlus == 1 else dfUnique_topicScoreS.sort_values(topicName)
    # display('dfUnique_topicScoreS:', dfUnique_topicScoreS) # для отладки
    
    docS_topic_pole = dfUnique_topicScoreS[dfUnique_topicScoreS[topicName] * minusPlus > 0]
    docS_topic_pole = docS_topic_pole.iloc[:min(docsLimit, len(docS_topic_pole)), :]
    # display('docS_topic_pole:', docS_topic_pole) # для supplementary

    return docS_topic_pole

# 1.2 ..описания каждого топика через его полюса и формирующие их токены и релевантные фрагменты располагаемых на них документов
def snippetByDoc(docS_topic_pole, loadingsThreshold, message_tokens_0, message_tokens_1, pole, supplementarieS, textFull_lemmatized, textFull_simbolsCleaned, tokenS_topic_pole_inUse_list):
    if pole != None: print(f'\n{pole.upper()}ЫЙ полюс топика')
    doc_snippetS = pandas.DataFrame(columns=['min', 'max', 'token'])
    doc_snippetS_new = pandas.DataFrame(columns=['min', 'max', 'token'])
    docs_snippetS = pandas.DataFrame(columns=['min', 'max', 'token'])
    # print('tokenS_topic_pole_inUse_list:', tokenS_topic_pole_inUse_list) # для отладки    
    if len(tokenS_topic_pole_inUse_list) == 0:
        if (loadingsThreshold != None) & (pole != None):
            print(f'Величина loadings токенов {pole}ого полюса не достигает заданного порога |{round(loadingsThreshold, 2)}|.'
                  , f'Поэтому {pole}ый полюс НЕ выражен и НЕ требует интерпретации')
    else:
        docS_topic_pole_list = list(docS_topic_pole.index)
        # print('docS_topic_pole_list:', docS_topic_pole_list) # для отладки

        print(f"{f'{pole.capitalize()}ый полюс' if pole != None else 'Топик'} сформирован токен{'ами' if len(tokenS_topic_pole_inUse_list) > 1 else 'ом'}",
              str(tokenS_topic_pole_inUse_list).replace('[', '').replace(']', ''), '-- по величине вклада')
        if message_tokens_1 != '': print(message_tokens_1)
        print(f"На {pole}ом полюсе расположен{'ы' if len(docS_topic_pole_list) > 1 else ''} следующи{'е' if len(docS_topic_pole_list) > 1 else 'й'} документ{'ы' if len(docS_topic_pole_list) > 1 else ''}" if pole != None else f"Топик выражен следующим{'и' if len(docS_topic_pole_list) > 1 else ''} документ{'ами' if len(docS_topic_pole_list) > 1 else 'ом'}",
              str(list(docS_topic_pole_list)).replace('[', '').replace(']', ''), '-- по близости к полюсу' if pole != None else '-- по степени вероятности')
        # display('docS_topic_pole:', docS_topic_pole) # для отладки

        for row in docS_topic_pole.index:
            indecesDuplicate_cellContent = docS_topic_pole['indicesDuplicate'][row]
            if len(indecesDuplicate_cellContent) > 0:
                print(f'На {pole}ом полюсе очищенный и лемматизированный текст документа' if pole != None else 'Текст документа', row,
                      f"дублируется в документ{'ах' if len(indecesDuplicate_cellContent) > 1 else 'е'}:",
                      str(list(indecesDuplicate_cellContent)).replace('[', '').replace(']', ''),
                      '-- поэтому далее не вывожу дубли')
        
        for doc in docS_topic_pole_list:
            # print(f'\nПредварительные фрагменты документа {doc}, содержащие указанные выше токены и их окружение.', 'Документ:', doc) # для отладки
            row = 0
            doc_snippetS = pandas.DataFrame(columns=['min', 'max', 'token'])
            # Предварительный проход по всем токенам на обрабатываемом полюсе для формирования doc_snippetS
            for token in tokenS_topic_pole_inUse_list:
                goC = True
                # print('Токен:', token) # для отладки
                textFull_lemmatized_list = docS_topic_pole[textFull_lemmatized][doc].split()
                textFull_list = docS_topic_pole[textFull_simbolsCleaned][doc].split()
                tokenProximity = 7  # лимит на длину окружения интересующего токена влево и вправо в нелемматизиованном документе
                tokenPosition = -1
                tokenPositionIntervalInitial = [0, 0]
                while goC:
                    try:
                        # Выяснить номер интересующего токена в списке всех токенов лемматизиованного документа
                        tokenPosition = textFull_lemmatized_list.index(token, tokenPosition + 1)

                        tokenPositionInterval = [tokenPosition - tokenProximity, tokenPosition + tokenProximity]
                        # print('tokenPositionIntervalInitial', tokenPositionIntervalInitial) # для отладки
                        # print('tokenPositionInterval', tokenPositionInterval) # для отладки

                        if tokenPositionIntervalInitial[-1] > tokenPositionInterval[0]:
                            # print('tokenPositionIntervalInitial и tokenPositionInterval пересеклись\n'
                            #       , 'Поэтому объединяю их\n'
                            #       , 'И не вывожу промежуточный фрагмент на экран')
                            tokenPositionInterval[0] = tokenPositionIntervalInitial[0]
                            # print('Объединённый tokenPositionInterval', tokenPositionInterval)
                        doc_snippetS.loc[row, 'min'] = tokenPositionInterval[0]
                        doc_snippetS.loc[row, 'max'] = tokenPositionInterval[-1]
                        doc_snippetS.loc[row, 'token'] = token
                        tokenPositionIntervalInitial = tokenPositionInterval.copy()
                        row += 1

                    except ValueError:
                        # print('Поиск в документе', doc, 'по токену "', token, '"завершён безуспешно') # для отладки
                        goC = False
                # display('doc_snippetS:', doc_snippetS) # для отладки
                
                doc_snippetS = doc_snippetS.drop_duplicates('min', keep='last')
                # display('doc_snippetS:', doc_snippetS) # для отладки

            # display('doc_snippetS:', doc_snippetS) # для отладки
            # Предварительный проход по всем токенам на обрабатываемом полюсе ЗАВЕРШЁН

            if len(doc_snippetS) > 0:
                # Проход по всем токенам на обрабатываемом полюсе для объединения пересекающихся фрагментов из doc_snippetS
                while True:
                    rowInvaderToDropS = []
                    # display('tokenS_topic_pole_inUse_list:', tokenS_topic_pole_inUse_list) # для отладки
                    tokenReceiverS = tokenS_topic_pole_inUse_list.copy()
                    for tokenInvader in tokenS_topic_pole_inUse_list[:-1]:
                        # display("doc_snippetS[doc_snippetS['token'] == tokenInvader]:", doc_snippetS[doc_snippetS['token'] == tokenInvader]) # для отладки                    
                        tokenInvaderIndeces = doc_snippetS[doc_snippetS['token'] == tokenInvader].index
                        # print('tokenInvaderIndeces:', tokenInvaderIndeces) # для отладки
                        tokenReceiverS.remove(tokenInvader)
                        for tokenReceiver in tokenReceiverS:
                            # print('tokenInvader:', tokenInvader, '-> tokenReceiver:', tokenReceiver) # для отладки
                            tokenReceiverIndeces = doc_snippetS[doc_snippetS['token'] == tokenReceiver].index
                            for rowInvader in tokenInvaderIndeces:
                                for rowReceiver in tokenReceiverIndeces:
                                    # print('Иду по:', rowInvader, rowReceiver) # для отладки
                                    # Сравнение диапазонов в строчках rowInvader и rowReceiver датафрейма doc_snippetS
                                    if (doc_snippetS['min'][rowInvader] >= doc_snippetS['min'][rowReceiver])\
                                            & (doc_snippetS['min'][rowInvader] <= doc_snippetS['max'][rowReceiver])\
                                        | (doc_snippetS['max'][rowInvader] >= doc_snippetS['min'][rowReceiver])\
                                            & (doc_snippetS['max'][rowInvader] <= doc_snippetS['max'][rowReceiver])\
                                        | (doc_snippetS['min'][rowReceiver] >= doc_snippetS['min'][rowInvader])\
                                            & (doc_snippetS['min'][rowReceiver] <= doc_snippetS['max'][rowInvader])\
                                        | (doc_snippetS['max'][rowReceiver] >= doc_snippetS['min'][rowInvader])\
                                            & (doc_snippetS['max'][rowReceiver] <= doc_snippetS['max'][rowInvader]):
                                        # print(rowInvader, rowReceiver)
                                        # display("rowInvader и rowReceiver concat:", pandas.concat([doc_snippetS.loc[rowInvader, :], doc_snippetS.loc[rowReceiver, :]])) # для отладки                    
                                        doc_snippetS.loc[rowReceiver, 'min'] = min(doc_snippetS['min'][rowInvader], doc_snippetS['min'][rowReceiver])
                                        doc_snippetS.loc[rowReceiver, 'max'] = max(doc_snippetS['max'][rowInvader], doc_snippetS['max'][rowReceiver])
                                        doc_snippetS.loc[rowReceiver, 'token'] += ' ' + tokenInvader
                                        # display('doc_snippetS:', doc_snippetS) # для отладки                    
                                        rowInvaderToDropS.append(rowInvader)
                    doc_snippetS = doc_snippetS.drop(rowInvaderToDropS)
                    # display('doc_snippetS на выходе из for in:', doc_snippetS) # для отладки
                    if (len(doc_snippetS) == len(doc_snippetS_new)) | (len(doc_snippetS) < 2): break # выход из while , т.к. на прошедшей итерации while ни один диапазон не объединился или нечего объединять 
                    else: # подготовка к следующей итерации while , чтобы попробовать объединить имеющиеся диапазоны  
                        doc_snippetS_new = doc_snippetS.copy()
                        tokenS_topic_pole_inUse_list = doc_snippetS['token'].tolist()
                # display('doc_snippetS на выходе из while:', doc_snippetS) # для отладки
                doc_snippetS['token'] = doc_snippetS['token'].apply(lambda cellContent:  ' '.join(sorted(set(cellContent.split())))) # удалить дубли токенов внутри каждой ячейки
                # display('doc_snippetS итоговый:', doc_snippetS) # для отладки
                # Проход по всем токенам на обрабатываемом полюсе для объединения пересекающихся фрагментов из doc_snippetS ЗАВЕРШЁН

                if len(doc_snippetS) > 0:
                    print(f'\nПосмотрите на фрагмент{'ы' if len(doc_snippetS) > 0 else ''} документа {doc}, содержащие указанные выше токены и их окружение.') # , 'Документ:', doc
                    for row in doc_snippetS.index:
                        # По границам интервала вывести окружение интересующего токена в нелемматизиованном документе
                        if pole != None: doc_snippetS.loc[row, 'pole'] = pole.capitalize() + 'ый'
                        doc_snippetS.loc[row, 'textSnippet'] = '..' + ' '.join(textFull_list[doc_snippetS['min'][row]: doc_snippetS['max'][row]]) + '..'
    
                        # print('Отладка')
                        # display('doc_snippetS до:', doc_snippetS)
                        # from openpyxl import Workbook
                        # wb = Workbook()
                        # ws = wb.create_sheet(title='topicName')
                        # for r in dataframe_to_rows(doc_snippetS, index=False, header=True):
                        #     ws.append(r)
    
                        doc_snippetS = supplementariesExecuter(docS_topic_pole, doc_snippetS, doc, row, supplementarieS)
    
                        # print('Отладка')
                        # display('doc_snippetS после:', doc_snippetS)
                        # from openpyxl import Workbook
                        # wb = Workbook()
                        # ws = wb.create_sheet(title='topicName')
                        # for r in dataframe_to_rows(doc_snippetS, index=False, header=True):
                        #     ws.append(r)
    
                        print(doc_snippetS['textSnippet'][row])

                docs_snippetS = pandas.concat([docs_snippetS, doc_snippetS])
                
                # print('Отладка')
                # display(docs_snippetS)
                # from openpyxl import Workbook
                # wb = Workbook()
                # ws = wb.create_sheet(title='topicName')
                # for r in dataframe_to_rows(docs_snippetS, index=False, header=True):
                #     ws.append(r)

            # doc_snippetS_all = pandas.concat([doc_snippetS_all, doc_snippetS])

        if len(docs_snippetS) == 0: print(message_tokens_0)
        print('\n')
        # display('Итоговый docs_snippetS в рамках функции:', docs_snippetS) # для отладки
    return docs_snippetS

# 1.3 ..оформления токенов на полюсах топиков
def summaryPole(minusPlus, tokenS_topic_pole_inUse_list, docS_topic, topicLoadingS, topicName):
    summaryPole = pandas.DataFrame()
    if len(tokenS_topic_pole_inUse_list) > 0:
        summaryPole = docS_topic[docS_topic[topicName] * minusPlus > 0].round(3)      
        summaryPole.loc[:, 'Топики'] = topicName
        summaryPole.loc[:, 'Токены'] = ', '.join(tokenS_topic_pole_inUse_list)
        # print('type(topicLoadingS):', type(topicLoadingS)) # для отладки
        if topicLoadingS is not None: summaryPole.loc[:, 'Усреднённая связь токена с топиком'] = round(topicLoadingS.loc[tokenS_topic_pole_inUse_list, topicName].mean(), 3)
        summaryPole.loc[:, 'Релевантность теме исследования'] = ''
        summaryPole.loc[:, 'Интерпретация топика'] = ''
    return summaryPole

# 1.4 ..внедрения вспомогательных полей (supplementaries) в итоговые датафреймы
def supplementariesExecuter(dfOriginator, dfRecipient, docIndex, dfRecipient_row, supplementarieS):
    if supplementarieS != None:
        # display('dfRecipient:', dfRecipient) # для отладки
        for supplementary in supplementarieS:
            try: dfRecipient.loc[dfRecipient_row, supplementary] = dfOriginator[supplementary][docIndex]
            except:
                # print(sys.exc_info()[1]) # для отладки
                # print('supplementary:', supplementary) # для отладки
                # print('dfRecipient.columns:', dfRecipient.columns) # для отладки
                dfOriginator[supplementary] = dfOriginator[supplementary].astype(str)
                dfRecipient.loc[dfRecipient_row, supplementary] = dfOriginator[supplementary][docIndex]
        return dfRecipient

# 1.5 ..выбора датафрейма и списка ключевых токенов полюса топика
def tokensSelector(docS_topic_pole, loadingsThreshold, minusPlus, textFull_lemmatized, tokensLimit, topicLoadingS, topicName):
    message_tokens_0 = ''
    message_tokens_1 = ''
    tokenS_topic = topicLoadingS.sort_values(topicName, ascending=False) if minusPlus == 1 else topicLoadingS.sort_values(topicName)
    # display('tokenS_topic.head(25):', tokenS_topic.head(25)) # для отладки
    
    tokenS_topic = tokenS_topic[tokenS_topic[topicName] * minusPlus > 0]
    tokenS_topic_list = tokenS_topic.index
    # print('tokenS_topic_list:', tokenS_topic_list) # для отладки

    # print('loadingsThreshold:', loadingsThreshold) # для отладки
    tokenS_topic_pole_inUse = tokenS_topic[tokenS_topic[topicName] * minusPlus > loadingsThreshold]
    tokenS_topic_pole_inUse = tokenS_topic_pole_inUse.iloc[:min(tokensLimit, len(tokenS_topic_pole_inUse)), :]
    tokenS_topic_pole_inUse_list = list(tokenS_topic_pole_inUse.index)
    # print('tokenS_topic_pole_inUse_list:', tokenS_topic_pole_inUse_list) # для отладки

    if len(tokenS_topic_pole_inUse_list) > 0: # если хотя бы один токен преодолевает порог loadingsThreshold

        docS_topic_pole_list = docS_topic_pole[textFull_lemmatized].tolist()
        docS_topic_pole_list = ' '.join(docS_topic_pole_list)
        docS_topic_pole_list = docS_topic_pole_list.split(' ')
        # print('docS_topic_pole_list:', docS_topic_pole_list) # для отладки

        goC = True
        while goC: # Определение списка ключевых токенов с учётом необходимости встречаемости хотя бы одного из них в ключевых документах того же полюса
            for token in tokenS_topic_pole_inUse_list:
                if token in docS_topic_pole_list:
                    goC = False
                    # print('Ключевой токен найден в хотя бы одном ключевом документе того же полюса') # для отладки
                    break
                else: message_tokens_1 = 'Порог tokensLimit был смягчён для обеспечения встречаемости хотя бы одного ключевого токена в ключевых документах того же полюса'
            if goC == True:
                tokensLimit += 1
                tokenS_topic_pole_inUse_new = tokenS_topic.iloc[:min(tokensLimit, len(tokenS_topic)), :] # добавить один ключевой токен полюса
                tokenS_topic_pole_inUse_new_list = list(tokenS_topic_pole_inUse_new.index)
                # print('tokenS_topic_pole_inUse_new_list:', tokenS_topic_pole_inUse_new_list) # для отладки
                if len(tokenS_topic_pole_inUse_list) == len(tokenS_topic_pole_inUse_new_list):
                    # если добавка ключевого токена полюса невозможна в силу исчерпания ключевых токенов полюса или достижения предела по loadingsThreshold
                    message_tokens_0 += '''В ключевых документах этого полюса не встречаются ключевые токены этого полюса. Возможная причина: ключевых токенов слишком мало в силу строгости порога loadingsThreshold или высокого порога tokensLimit.
--- Если хотите получить фрагменты ключевых документов, относящихся к ключевым токенам, попробуйте снизить перечисленные пороги и перезапустить функцию randanTopic .'''
                    goC = False
                else: tokenS_topic_pole_inUse_list = tokenS_topic_pole_inUse_new_list # подготовка новой итерации после добавки одного ключевого токена полюса
    # print('message_tokens:', message_tokens) # для отладки
    return message_tokens_0, message_tokens_1, tokenS_topic_pole_inUse, tokenS_topic_pole_inUse_list

def randanTopic(dfIn, matrix_df, docsLimit=5, loadingsThreshold=0.5, returnDfs=False, rowsNumerator=None, supplementarieS=None, textFull_lemmatized='textFull_lemmatized', textFull_simbolsCleaned='textFull_simbolsCleaned', tokensLimit=10, topicsCount=None):
    '''    Метод тематического моделирования randanTopic основан на методе главных компонент (по-английски: principal components analisys, PCA). То есть он НЕ использует нейросети и embeddings, а работает с "мешком слов" (по-английски: bag of words, BoW). Поэтому важно качественно подготовить "мешок слов" через очистку исходного текста от лишних символов, лемматизацию и удаление стоп-слов. Эти три этапа предобработки исходного текста, а также (при необходимости) автокоррекция грамматических ошибок могут быть выполнены функциями из скрипта textPreprocessor пакета randan https://github.com/RandanCSS/randan/blob/master/randan/tools/textPreprocessor.py . Причём для удобства последующей интерпретации рекомендую результаты очистки от лишних символов и результаты лемматизации сохранить в отдельные столбцы: textFull_simbolsCleaned и textFull_lemmatized соответственно.
        Если для последующей интерпретации Вам пригодятся дополнительные столбцы (скажем, заголовок, дата и т.п.), впишите их в формате списка текстовых объектов в аргумент supplementaries= текущей функции.
    
    Parameters
    --------------------
                   dfIn : DataFrame -- таблица с исходными данными (текстами и, при ихналичии, вспомогательными переменными)
              matrix_df : DataFrame -- мешок слов
      loadingsThreshold : float -- порог (по модулю) усреднённой связи токена с топиком; нужен для отбора токенов, косвенно лимитирует число токенов при интерпретации         
              docsLimit : int -- лимит на число документов на полюсе топика; нужен для отбора документов            
              returnDfs : bool -- в случае True функция возвращает датафреймы с фрагментами документов docs_snippetS и реквизитами документов и вспомогательными переменными scoreS
          rowsNumerator : str -- имя столбца, значения которого будут использоваться в качестве обозначения строк итоговых таблиц
        supplementarieS : list -- дополнительные столбцы для последующей интерпретации
    textFull_lemmatized : str -- имя столбца с текстом, прошедшем лемматизацию, но из которого НЕ удалены стоп-слова
textFull_simbolsCleaned : str -- имя столбца с текстом, прошедшем удаление лишних символов, но НЕ прошедший лемматизацию и из которого НЕ удалены стоп-слова
            tokensLimit : int -- лимит на число ключевых токенов на полюсе топика. Этот порог может быть автоматически смягчён в процессе исполнения алгоритма для обеспечения встречаемости хотя бы одного ключевого токена в ключевых документах того же полюса
            topicsCount : int -- частота самого высокочастотного слова из тех, которые предложены автокорректором в качестве правильного варианта; этот аргумент необходим для корректной работы аргумента userWordS'''

    df = dfIn.copy()

    if (returnDfs == False) & (rowsNumerator == None) & (supplementarieS == None) & (topicsCount == None):
        # print('Пользователь не подал аргументы') # для отладки
        expiriencedMode = False
    else:
        expiriencedMode = True

    if expiriencedMode == False:
        print(
'''    Метод тематического моделирования randanTopic основан на методе главных компонент (по-английски: principal components analisys, PCA). То есть он НЕ использует нейросети и embeddings, а работает с "мешком слов" (по-английски: bag of words, BoW). Поэтому важно качественно подготовить "мешок слов" через очистку исходного текста от лишних символов, лемматизацию и удаление стоп-слов. Эти три этапа предобработки исходного текста, а также (при необходимости) автокоррекция грамматических ошибок могут быть выполнены функциями из скрипта textPreprocessor пакета randan https://github.com/RandanCSS/randan/blob/master/randan/tools/textPreprocessor.py . Причём для удобства последующей интерпретации рекомендую результаты очистки от лишних символов и результаты лемматизации сохранить в отдельные столбцы: textFull_simbolsCleaned и textFull_lemmatized соответственно.
--- После прочтения этой инструкции нажмите Enter'''
              )
        input()

    # print('textFull_lemmatized', textFull_lemmatized) # для отладки
    # print('textFull_lemmatized in df.columns', textFull_lemmatized in df.columns) # для отладки
    textFull_lemmatized = columnTargetChecker(textFull_lemmatized, 'Какой столбец в поданном Вами датафрейме содержит лемматизированные тексты (из которого НЕ удалены стоп-слова)?', df)
#     if textFull_lemmatized not in df.columns:
#         print(
# f'''В поданном Вами датафрейме отсутствует столбец '{textFull_lemmatized}' . Какой столбец в поданном Вами датафрейме содержит лемматизированные тексты (из которого НЕ удалены стоп-слова)?'''
#               )
#         while True:
#             print('--- Впишите, пожалуйста, его название')
#             textFull_lemmatized = input()
#             if textFull_lemmatized in df.columns: break
#             else:
#                 print('--- Вы вписали что-то не то')

    # print('textFull_simbolsCleaned', textFull_simbolsCleaned) # для отладки
    # print('textFull_simbolsCleaned in df.columns', textFull_simbolsCleaned in df.columns) # для отладки
    textFull_simbolsCleaned = columnTargetChecker(textFull_simbolsCleaned, 'Какой столбец в поданном Вами датафрейме содержит тексты, из которых удалены лишние символы, но НЕ лемматизированные?', df)
#     if textFull_simbolsCleaned not in df.columns:
#         print(
# f'''В поданном Вами датафрейме отсутствует столбец '{textFull_simbolsCleaned}' . Какой столбец в поданном Вами датафрейме содержит тексты, из которых удалены лишние символы, но НЕ лемматизированные?'''
#               )
#         while True:
#             print('--- Впишите, пожалуйста, его название')
#             textFull_lemmatized = input()
#             if textFull_lemmatized in df.columns: break
#             else:
#                 print('--- Вы вписали что-то не то')  
                    
# 0. Определение желаемого числа топиков
    if topicsCount == None:
        print(
'''Вы уже знаете, сколько топиков хотите, или предпочитаете посмотреть на возможные ориентиры числа топиков, чтобы затем выбрать их число?
--- Если знаете, то введите желаемое число топиков и нажмите Enter
--- Если предпочитаете посмотреть, то просто нажмите Enter -- по умолчанию будут выведены ориентиры для числа топиков не более 30,
поскольку большее число топиков (а) вряд ли захотите проинтерпретировать и (б) расчёт ориентиров для них может занять слишком продолжительное время'''
              )
        while True:
            topicsCount = input()
            try:
                topicsCount = None if topicsCount == '' else int(topicsCount)
                break
            except ValueError:
                print('--- Вы вписали что-то не то'
                      , '\n--- Впишите, пожалуйста, целое число и нажмите Enter: ')

        if topicsCount == None:
            print(
'''Ориентиры для выбора числа топиков: (а) описательная статистика частотности токенов во всём корпусе и в отдельных документах, (б) приемлемый уровнь объяснительной способности модели, критерий Кайзера и скачки объяснительной способности модели (диаграмма "каменистой осыпи"), (в) Ваша субъективная готовность проинтерпретировать не более стольки-то топиков.'''
                  )
# 1.0. Первичный анализ частот токенов в корпусе документов в целом и по каждому документу
# Посчитать сумму по ТОКЕНАМ
            встречаемостьТокенов = pandas.DataFrame(matrix_df.sum(), columns=['Встречаемость токенов'])
            # display(встречаемостьТокенов)
# Посчитать сумму по ДОКУМЕНТАМ
            токеныДокумент = pandas.DataFrame(matrix_df.T.sum(), columns=['Наполненность токенами документов'])
            # display(токеныДокумент)
# 1.1. Описательная статистика для интервальной переменной встречаемостьТокенов
            descriptive_statistics.ScaleStatistics(встречаемостьТокенов[['Встречаемость токенов']])
    # .. и для интервальной переменной токеныДокумент
            descriptive_statistics.ScaleStatistics(токеныДокумент[['Наполненность токенами документов']])

# 2. Пробный запуск (без вращения, но с выводом результатов), чтобы посмотреть `Explained variance`
            dimension_reduction.PCA(n_components='Kaiser' if len(matrix_df.columns) < 30 else 30).fit(matrix_df)
            print('\n--- Введите желаемое число топиков и нажмите Enter')
            while True:
                topicsCount = input()
                try:
                    topicsCount = int(topicsCount)
                    break
                except ValueError:
                    print('--- Вы вписали что-то не то'
                          , '\n--- Впишите, пожалуйста, целое число и нажмите Enter: ')

# 3. Итоговый запуск (с вращением, но с без вывода части результатов)
# Настроить класс PCA: предупредить, сколько будет топиков и какое будет вращение
    if topicsCount > len(matrix_df.columns):
        print('Число топиков принудительно снижено до', len(matrix_df.columns), ', поскольку значение topicsCount, равное', topicsCount, ', слишком велико для располагаемых данных.')
        topicsCount = len(matrix_df.columns)
    pca = dimension_reduction.PCA(n_components=min(len(matrix_df.columns), 300), rotation='varimax')
# Подать токены в настроенный класс PCA
    pca = pca.fit(matrix_df, show_results=False)

# 4. Четыре датафрейма, ключевых для оформления и интерпретации результатов тематического моделирования, плюс один датафрейм
    component_loadings_rotated = pca.component_loadings_rotated
    component_loadings_rotated = component_loadings_rotated.iloc[:, :topicsCount]
    display(component_loadings_rotated) # для отладки
    topicNameS = component_loadings_rotated.columns

# Матрица "документы-топики" (и величины в ячейках матрицы названы так же)
    scoreS = pca.transform(matrix_df)
    scoreS = scoreS.iloc[:, :topicsCount]
    # display(scoreS) # для отладки	
    
    if rowsNumerator != None:
        if (sum(df[rowsNumerator].isna()) == 0) & (len(list(set(df[rowsNumerator].tolist()))) == len(df)): df.index = df[rowsNumerator] # значения назначенного пользователем столбца будут использоваться в качестве обозначения строк итоговых таблиц, только если эти значения без NaN и без дубликатов
        else: print(
'''
Назначенный Вами столбец для обозначения строк итоговых таблиц содержит NaN или дубликаты, поэтому не будет использоваться.
'''
                    )
    else: # в противном случае останутся исходные обозначения строк, но если эти значения без дубликатов -- иначе reset_index
        if len(list(set(list(df.index)))) != len(df): df.reset_index(drop=True)
        print(
'''
Cреди обозначений строк исходной таблицы есть дубликаты, поэтому эти обозначения заменяю на сквозную нумерацию.
'''
              )

# Добавить столбец со списками дублирующих (по столбцу textFull_lemmatized) каждый текст текстов и убрать неиспользуемые столбцы
    df_columnS = [textFull_lemmatized, textFull_simbolsCleaned]
    if supplementarieS != None: df_columnS.extend(supplementarieS)
    df = df[df_columnS]
                               
    # Добавить столбец со списком всех номеров строк для каждого текста
    df['indicesDuplicate'] = df.groupby(textFull_lemmatized)[textFull_lemmatized].transform(lambda group: [group.index.tolist()] * len(group)) # 'textFull_stopwordsDropped'

    # Убрать собственный индекс из списка дубликатов в каждой оставшейся строке
    df['indicesDuplicate'] = df.apply(lambda row: [i for i in row['indicesDuplicate'] if i != row.name], axis=1)
    # display('df:', df) # для отладки

    scoreS.index = df.index # связать документы и scoreS

# 5. Цикл для прохода по всем топикам
    poleS = ['лев', 'прав']
    
    summary = pandas.DataFrame()
    docs_snippetS = pandas.DataFrame()
    errorS = []

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Сводка"

    for topicName in topicNameS[:topicsCount]:
    # for topicName in topicNameS[2:4]: # для отладки
        print('\n\nTopic', topicName)

        # Документы-топики
        topicScoreS = scoreS[[topicName]]

        # Токены-топики
        topicLoadingS = component_loadings_rotated[[topicName]]
        # display(topicLoadingS) # для отладки

        # Максимум (по модулю) нагрузки внутри компоненты
        maxLoading = topicLoadingS[topicName].abs().max()
        # print('maxLoading:', maxLoading) # для отладки
        loadingsThreshold_user = loadingsThreshold # чтобы по истечении итерации вернуться к заданному пользователем значени		
        if loadingsThreshold > maxLoading:
            print('Порог (по модулю) усреднённой связи токена с топиком снижен до', round(maxLoading, 2), ', поскольку значение loadingsThreshold, равное', round(loadingsThreshold, 2), ', слишком велико для располагаемых данных.')
            loadingsThreshold = maxLoading * 0.99 # не выше минимума из максимальных по модулю loadings среди компонент			

        plt.figure(figsize=(9, 9))

    # Распределение документов по осям
        # Гистограмма
        plt.subplot(3, 2, 1) # три строки, три столбца, первое место
        plt.hist(topicScoreS, color='grey')
        plt.title(f"Docs Distribution by Scores\nin {topicName}")
        plt.xlabel('Scores')
        plt.ylabel('Frequency');

        # Боксплот
        plt.subplot(3, 2, 2) # три строки, три столбца, третье место
        plt.boxplot(topicScoreS)
        plt.title(f"Docs Distribution by Scores\nin {topicName}")
        plt.xticks([])
        plt.ylabel('Scores')
        # plt.show()

    # Распределение токенов по осям
        # Гистограмма
        plt.subplot(3, 2, 3) # три строки, три столбца, седьмое место
        plt.hist(topicLoadingS.dropna(), color='grey')
        plt.title(f"Tokens Distribution by Loadings\nin {topicName}")
        plt.xlabel('Loadings')
        plt.ylabel('Frequency');

        # Боксплот
        plt.subplot(3, 2, 4) # три строки, три столбца, девятое место
        plt.boxplot(topicLoadingS.dropna())
        plt.title(f"Tokens Distribution by Loadings\nin {topicName}")
        plt.xticks([])
        plt.ylabel('Loadings')
        
        plt.tight_layout()
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        plt.show() # обязательно после savefig

        # Полярные документы, причём уникальные

        # Добавить столбец со scoreS документов в рамках рассматриваемого топика
        # display('topicScoreS:', topicScoreS) # для отладки        
        df_topicScoreS = pandas.concat([df, topicScoreS], axis=1)
    
        # Оставить только уникальные (по столбцу textFull_lemmatized) тексты (первые вхождения)
        dfUnique_topicScoreS = df_topicScoreS.drop_duplicates(textFull_lemmatized, keep='first') # 'textFull_stopwordsDropped'
    
        # Убрать собственный индекс из списка дубликатов в каждой оставшейся строке
        dfUnique_topicScoreS['indicesDuplicate'] = dfUnique_topicScoreS.apply(lambda row: [i for i in row['indicesDuplicate'] if i != row.name], axis=1)
        # display('dfUnique_topicScoreS:', dfUnique_topicScoreS) # для отладки
    
        docS_topic_minus = docsSelector(dfUnique_topicScoreS, -1, docsLimit, topicName)
        docS_topic_plus = docsSelector(dfUnique_topicScoreS, 1, docsLimit, topicName)
        
        # Полярные токены
        message_tokens_0, message_tokens_1, tokenS_topic_minus_inUse, tokenS_topic_minus_inUse_list = tokensSelector(docS_topic_minus, loadingsThreshold, -1, textFull_lemmatized, tokensLimit, topicLoadingS, topicName)
        message_tokens_0, message_tokens_1, tokenS_topic_plus_inUse, tokenS_topic_plus_inUse_list = tokensSelector(docS_topic_plus, loadingsThreshold, 1, textFull_lemmatized, tokensLimit, topicLoadingS, topicName)

        print('Токены на полюсах топика', topicName)
        display(pandas.concat([tokenS_topic_minus_inUse, tokenS_topic_plus_inUse])) #.dropna()
        
        # Описание каждого топика через его полюса и формирующие их токены и релевантные фрагменты располагаемых на них документов
        # display('topicLoadingS:', topicLoadingS) # для отладки

        # print('tokenS_topic_minus_inUse_list:', tokenS_topic_minus_inUse_list) # для отладки
        # display('Средняя связь ключевых токенов с топиком:', tokenS_topic_minus_inUse[topicName].mean()) # для отладки

        # print('docS_topic_minus_list:', docS_topic_minus_list[:2]) # для отладки
        docs_snippetS_minus = snippetByDoc(docS_topic_minus, loadingsThreshold, message_tokens_0, message_tokens_1, poleS[0], supplementarieS, textFull_lemmatized, textFull_simbolsCleaned, tokenS_topic_minus_inUse_list)
        
        # print('tokenS_topic_plus_inUse_list:', tokenS_topic_plus_inUse_list) # для отладки
        # display('Средняя связь ключевых токенов с топиком:', tokenS_topic_plus_inUse[topicName].mean()) # для отладки

        # print('docS_topic_plus_list:', docS_topic_plus_list[:2]) # для отладки
        docs_snippetS_plus = snippetByDoc(docS_topic_plus, loadingsThreshold, message_tokens_0, message_tokens_1, poleS[-1], supplementarieS, textFull_lemmatized, textFull_simbolsCleaned, tokenS_topic_plus_inUse_list)
        
        # print('Отладка')
        # display(docs_snippetS_minus, docs_snippetS_plus)
        # from openpyxl import Workbook
        # wb = Workbook()
        # ws = wb.create_sheet(title='topicName')
        # for r in dataframe_to_rows(docs_snippetS_minus, index=False, header=True):
        #     ws.append(r)
        # for r in dataframe_to_rows(docs_snippetS_plus, index=False, header=True):
        #     ws.append(r)

        docs_snippetS_additional = pandas.concat([docs_snippetS_minus, docs_snippetS_plus])
        # display('docs_snippetS_additional:', docs_snippetS_additional) # для отладки
        
        docs_snippetS_additional = docs_snippetS_additional.drop(['min', 'max'], axis=1)
        docs_snippetS_additional = docs_snippetS_additional.reset_index(drop=True)
        
    # Обработка полярных документов и токенов; запись в датафрейм
        docS_topic = pandas.concat([docS_topic_minus, docS_topic_plus])
        # display('docS_topic:', docS_topic) # для отладки

        # Заполнение docs_snippetS_additional , если он пустой
        if len(docs_snippetS_additional) == 0:
            docs_snippetS_additional = docS_topic[[textFull_simbolsCleaned]]
            if len(tokenS_topic_minus_inUse_list) > 0: docs_snippetS_additional.loc[docS_topic_minus.index, 'pole'] = poleS[0].capitalize() + 'ый'
            else: docs_snippetS_additional = docs_snippetS_additional.drop(docS_topic_minus.index) # чтобы не выводить в эксельке ключевые документы при отсутствии ключевых токенов

            if len(tokenS_topic_plus_inUse_list) > 0: docs_snippetS_additional.loc[docS_topic_plus.index, 'pole'] = poleS[-1].capitalize() + 'ый'
            else: docs_snippetS_additional = docs_snippetS_additional.drop(docS_topic_plus.index) # чтобы не выводить в эксельке ключевые документы при отсутствии ключевых токенов

            docs_snippetS_additional = docs_snippetS_additional[['pole', textFull_simbolsCleaned]]
            for doc in docs_snippetS_additional.index: # гипотетически, проблема невосприимчивости экселя к сложной структуре внутри ячеек решается , когда датафрейм проходит через for in
                # display('docs_snippetS_additional:', docs_snippetS_additional) # для отладки
                docs_snippetS_additional = supplementariesExecuter(df, docs_snippetS_additional, doc, doc, supplementarieS)
        
        # display('docs_snippetS_additional:', docs_snippetS_additional) # для отладки            
        docs_snippetS_additional.loc[:, 'Интерпретация топика'] = ''
        docs_snippetS = pandas.concat([docs_snippetS, docs_snippetS_additional])
        print('\n')
        
        summaryMinus = summaryPole(-1, tokenS_topic_minus_inUse_list, docS_topic, topicLoadingS, topicName)
        summaryPlus = summaryPole(1, tokenS_topic_plus_inUse_list, docS_topic, topicLoadingS, topicName)
        summary_additional = pandas.concat([summaryMinus, summaryPlus])
        # display('summary_additional:', summary_additional) # для отладки

        summary = pandas.concat([summary, summary_additional])
        summary['Документы'] = summary.index
        # display('summary:', summary) # для отладки

        # print('Отладка')
        # display(docs_snippetS_additional)
        # from openpyxl import Workbook
        # wb = Workbook()
        # ws = wb.create_sheet(title='topicName')
        # for r in dataframe_to_rows(docs_snippetS_additional, index=False, header=True):
        #     ws.append(r)        

        ws = wb.create_sheet(title=topicName)
        for r in dataframe_to_rows(docs_snippetS_additional, index=False, header=True):
            ws.append(r)
        
        # while True:
        #     try:
        #         ws = wb.create_sheet(title=topicName)
        #         for r in dataframe_to_rows(docs_snippetS_additional, index=False, header=True):
        #             ws.append(r)
        #             # r_previous = r
        #         break
        #     except:
        #         # ws.remove(r_previous)
        #         # print(sys.exc_info()[1]) # для отладки
        #         # docs_snippetS_additional['title'] = docs_snippetS_additional['title'].astype(str)
        #         docs_snippetS_additional = docs_snippetS_additional.astype(str)

        img_data.seek(0)
        img = Image(img_data)
        ws.add_image(img, "K1")
        loadingsThreshold = loadingsThreshold_user # возврат к заданному пользователем значению по истечении итерации

    docs_snippetS = docs_snippetS.reset_index(drop=True)

# 6. Если норм, отправить эти документы в Excel
    # Желательно после начала интерпретации назвать Excel по-новому, чтобы он не перезаписался в дальнейшем.
    summaryColumns = ['Топики', 'Токены', 'Усреднённая связь токена с топиком', 'Документы'] 
    summaryColumns.extend(scoreS.columns)
    summaryColumns.extend(['Релевантность теме исследования', 'Интерпретация топика'])

    for r in dataframe_to_rows(summary[summaryColumns], index=False, header=True):
        summary_ws.append(r)

    fileName = 'Топики'
    attempt = 0
# Чтобы не перезаписать ранее созднный файл, в котором может быть интерпретация
    while os.path.exists(fileName + ' ' + str(attempt) + ".xlsx"):
        attempt += 1

    print(f'''\nCоздаю файл "{fileName + ' ' + str(attempt)}.xlsx"''')
    wb.save(fileName + ' ' + str(attempt) + ".xlsx")

    print(
'''
Рекомендации по дальнейшей интерпретации топиков:
1.1. Интерпретация топиков по ключевым токенам: на листе Сводка посмотрите перечень ключевых токенов каждого топика. Если токенов слишком мало или много, перезапустите скрипт, понизив или повысив порог loadingsThreshold соответственно, либо повысив или понизив порог tokensLimit соответственно.
1.2. Попробуйте в одной или нескольких фразах сормулировать смысл каждого топика по его ключевым токенам и внесите сформулированный смысл в столбец Интерпретация топика. Если у топика два полюса, лучше для каждого из них формулировать смысл отдельно.

2.1. Интерпретация топиков по выдержкам из ключевых документов и ключевым токенам: на листе Сводка посмотрите перечень ключевых документов каждого топика. Если документов слишком мало или много, перезапустите скрипт, повысив или понизив порог docsLimit соответственно.
2.2. На листе PC1_vrmx читайте выдержки (с акцентом на соответствующие ключевые токены) и при обнаружении принципиально нового смысла, вносите его в свободный столбец. Нюансы и оттенки уже выписанных смыслов лучше игнорировать, чтобы не "закопаться".
2.3. Повторите процедуру для остальных листов PC..._vrmx. По завершении постарайтесь в одной или нескольких фразах обобщить смыслы для каждого листа (топика) и внесите сформулированные обобщения в столбец Интерпретация топика листа Сводка. Если у топика два полюса, лучше для каждого из них формулировать смысл отдельно.

3.1. Интерпретация топиков по ключевым документам и ключевым токенам с помощью ИИ: на листе Сводка посмотрите перечень ключевых токенов и документов каждого топика. Если их слишком мало или много, выполните пункты 2.1 или 3.1.
3.2. На листе Сводка выберите несколько (можете ориентироваться на диаграмму Docs distribution by scores...) ключевых документов с максимальным значением в столбце PC1_vrmx , подайте их в ИИ с просьбой вычленить основные мысли с привязкой к соответствующим ключевым токенам.
3.3. Ответ ИИ внесите в столбец Интерпретация топика напротив первого топика. Повторите процедуру для остальных топиков.'''
          )
    if returnDfs: return docs_snippetS, scoreS
